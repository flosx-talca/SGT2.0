from flask import Blueprint, request, jsonify, render_template
from flask_login import login_required, current_user
from sqlalchemy import select, desc
from app import db
from app.models.scheduling import CuadranteCabecera, CuadranteAsignacion
from app.models.business import Trabajador, Turno
from app.services.cuadrante_service import guardar_cuadrante, editar_asignacion_manual
from flask import send_file, Response, make_response
import calendar
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Flowable
from reportlab.lib.styles import getSampleStyleSheet
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

cuadrante_bp = Blueprint('cuadrante', __name__, url_prefix='/cuadrante')

@cuadrante_bp.route('/guardar', methods=['POST'])
@login_required
def guardar():
    """Guarda el cuadrante generado por el Solver."""
    data = request.get_json()
    ip = request.headers.get('X-Forwarded-For', request.remote_addr)
    cabecera = guardar_cuadrante(
        empresa_id=data['empresa_id'],
        servicio_id=data.get('servicio_id'),
        mes=data['mes'],
        anio=data['anio'],
        asignaciones=data['asignaciones'],
        ip=ip
    )
    return jsonify({"ok": True, "cabecera_id": cabecera.id})


@cuadrante_bp.route('/asignacion', methods=['PUT'])
@login_required
def editar_asignacion():
    """Modifica una asignación manualmente post-guardado."""
    data   = request.get_json()
    ip     = request.headers.get('X-Forwarded-For', request.remote_addr)
    
    try:
        asig = editar_asignacion_manual(
            cabecera_id=data.get('cabecera_id'),
            trabajador_id=data.get('trabajador_id'),
            fecha=data.get('fecha'),
            turno_nuevo_id=data.get('turno_id'),
            es_libre=data.get('es_libre', False),
            motivo=data.get('motivo', 'Cambio manual UI'),
            ip=ip
        )
        return jsonify({
            "ok": True,
            "asignacion_id": asig.id,
            "origen": asig.origen
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 400


@cuadrante_bp.route('/lista', methods=['GET'])
@login_required
def lista():
    """Retorna las últimas planificaciones para el DataTable del dashboard."""
    from app.services.context import get_empresas_usuario
    
    # 1. Obtener empresas permitidas
    empresas = get_empresas_usuario()
    ids_permitidos = [e.id for e in empresas]
    
    # 2. Consultar cabeceras filtradas
    query = select(CuadranteCabecera).where(CuadranteCabecera.empresa_id.in_(ids_permitidos))
    
    cabeceras = db.session.execute(
        query.order_by(desc(CuadranteCabecera.guardado_en)).limit(50)
    ).scalars().all()

    return render_template(
        'cuadrante/lista_partial.html',
        cabeceras=cabeceras
    )
@cuadrante_bp.route('/verificar', methods=['GET'])
@login_required
def verificar_existencia():
    """Verifica si ya existe un cuadrante para el período y servicio."""
    mes = request.args.get('mes', type=int)
    anio = request.args.get('anio', type=int)
    empresa_id = request.args.get('empresa_id', type=int)
    servicio_id = request.args.get('servicio_id', type=int)

    if not mes or not anio or not servicio_id:
        return jsonify({"existe": False, "error": "Faltan parámetros"}), 400

    query = select(CuadranteCabecera).where(
        CuadranteCabecera.servicio_id == servicio_id,
        CuadranteCabecera.mes == mes,
        CuadranteCabecera.anio == anio
    )
    
    if empresa_id:
        query = query.where(CuadranteCabecera.empresa_id == empresa_id)

    existente = db.session.execute(query).scalar_one_or_none()

    return jsonify({
        "existe": existente is not None,
        "cabecera_id": existente.id if existente else None
    })

@cuadrante_bp.route('/publicar', methods=['POST'])
@login_required
def publicar():
    """Marca el cuadrante como publicado (inmutable)."""
    data = request.get_json()
    cabecera_id = data.get('cabecera_id')
    
    cabecera = db.session.get(CuadranteCabecera, cabecera_id)
    if not cabecera:
        return jsonify({"ok": False, "message": "Cuadrante no encontrado"}), 404
        
    try:
        from datetime import datetime
        cabecera.estado = 'publicado'
        # Podríamos agregar campos de auditoría de publicación si existieran en el modelo
        # Por ahora usamos el estado para bloquear
        db.session.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "message": str(e)}), 400

@cuadrante_bp.route('/exportar/excel/<int:cabecera_id>')
@login_required
def exportar_excel(cabecera_id):
    """Generar Excel Vertical con Títulos y Nombres Completos."""
    try:
        cabecera = db.session.get(CuadranteCabecera, cabecera_id)
        if not cabecera:
            return "Cuadrante no encontrado", 404

        # 1. Datos y Mapeos
        asignaciones = CuadranteAsignacion.query.filter_by(cabecera_id=cabecera_id).all()
        trabajadores_db = Trabajador.query.filter_by(empresa_id=cabecera.empresa_id, servicio_id=cabecera.servicio_id, activo=True).all()
        turnos_db = Turno.query.filter_by(empresa_id=cabecera.empresa_id, activo=True).all()
        
        t_nombres = {t.id: f"{t.nombre} {t.apellido1}" for t in trabajadores_db}
        turnos_map = {t.id: t.abreviacion for t in turnos_db}
        colores_map = {t.abreviacion: t.color.replace('#', '') if t.color else 'FFFFFF' for t in turnos_db}
        duraciones_map = {t.id: t.duracion_hrs for t in turnos_db}
        
        # Feriados
        import calendar
        _, last_day = calendar.monthrange(cabecera.anio, cabecera.mes)
        from app.models.core import Feriado
        feriados = {f.fecha.day for f in Feriado.query.filter(
            db.extract('month', Feriado.fecha) == cabecera.mes,
            db.extract('year', Feriado.fecha) == cabecera.anio,
            Feriado.activo == True
        ).all()}

        # 2. Matriz de Datos
        dias = list(range(1, last_day + 1))
        t_ids = [t.id for t in trabajadores_db]
        dias_nombres = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
        
        matrix = []
        for d in dias:
            fecha_dt = datetime(cabecera.anio, cabecera.mes, d)
            nombre_dia = dias_nombres[fecha_dt.weekday()]
            row = {'Día': f"{nombre_dia} {d:02d}"}
            for tid in t_ids:
                asig = next((a for a in asignaciones if a.fecha.day == d and a.trabajador_id == tid), None)
                row[t_nombres[tid]] = turnos_map.get(asig.turno_id, "L") if (asig and not asig.es_libre) else "L"
            matrix.append(row)
        
        df = pd.DataFrame(matrix).set_index('Día')

        # 3. Construcción Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Planificación', startrow=2)
            workbook = writer.book
            worksheet = writer.sheets['Planificación']

            # Calcular ancho total exacto (Día + Trabajadores + Espacio + Resúmenes)
            total_cols = len(t_ids) + 2 + len(turnos_db)

            # Título
            titulo = f"Cuadrante {cabecera.empresa.razon_social} - {cabecera.servicio.descripcion} ({cabecera.periodo})"
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
            cell_title = worksheet.cell(row=1, column=1, value=titulo)
            cell_title.font = Font(bold=True, size=14)
            cell_title.alignment = Alignment(horizontal='center')

            # Estilos
            midnight = "1a2a3a"
            header_fill = PatternFill(start_color=midnight, end_color=midnight, fill_type="solid")
            holiday_fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            v_align = Alignment(textRotation=90, horizontal='center', vertical='center')
            c_align = Alignment(horizontal='center', vertical='center')
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Formatear Cabeceras (Fila 3)
            worksheet.row_dimensions[3].height = 110
            for col_idx in range(1, total_cols + 1):
                cell = worksheet.cell(row=3, column=col_idx)
                if col_idx != len(t_ids) + 2: # No pintar la columna de separación
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = v_align if col_idx > 1 else c_align
                    cell.border = border

            # 4. Datos y Resúmenes
            for idx, d in enumerate(dias):
                row_idx = idx + 4
                fecha_dt = datetime(cabecera.anio, cabecera.mes, d)
                is_h = d in feriados or fecha_dt.weekday() == 6
                
                # Día
                c_dia = worksheet.cell(row=row_idx, column=1)
                c_dia.border = border
                c_dia.alignment = c_align
                if is_h:
                    c_dia.fill = holiday_fill
                    c_dia.font = Font(color="FFFFFF", bold=True)
                
                # Turnos
                for c_idx in range(2, len(t_ids) + 2):
                    cell = worksheet.cell(row=row_idx, column=c_idx)
                    val = str(cell.value)
                    if val in colores_map:
                        cell.fill = PatternFill(start_color=colores_map[val], end_color=colores_map[val], fill_type="solid")
                        cell.font = Font(bold=True)
                    cell.border = border
                    cell.alignment = c_align

                # Resumen Diario (Derecha)
                col_res_start = len(t_ids) + 3
                for i, t_obj in enumerate(turnos_db):
                    count = (df.iloc[idx] == t_obj.abreviacion).sum()
                    r_cell = worksheet.cell(row=row_idx, column=col_res_start + i, value=count)
                    r_cell.border = border
                    r_cell.alignment = c_align
                    if row_idx == 4:
                        h_cell = worksheet.cell(row=3, column=col_res_start + i, value=f"Total {t_obj.nombre}")
                        h_cell.fill = header_fill
                        h_cell.font = header_font
                        h_cell.alignment = v_align
                        h_cell.border = border

            # 5. Totales Trabajador (Abajo)
            f_start = last_day + 4
            worksheet.cell(row=f_start, column=1, value="HORAS").font = Font(bold=True)
            worksheet.cell(row=f_start+1, column=1, value="TURNOS").font = Font(bold=True)
            for c_idx, tid in enumerate(t_ids):
                t_asigs = [a for a in asignaciones if a.trabajador_id == tid]
                hrs = sum(duraciones_map.get(a.turno_id, 0) for a in t_asigs if not a.es_libre)
                tns = sum(1 for a in t_asigs if not a.es_libre)
                c_h = worksheet.cell(row=f_start, column=c_idx + 2, value=hrs)
                c_t = worksheet.cell(row=f_start+1, column=c_idx + 2, value=tns)
                c_h.border = border
                c_t.border = border

            # Ajustes finales
            worksheet.column_dimensions['A'].width = 10
            for col_idx in range(2, total_cols + 1):
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 4.5

        output.seek(0)
        # Nombre de archivo dinámico: Empresa_Servicio_Mes_Anio.xlsx
        empresa_clean = cabecera.empresa.razon_social.replace(" ", "_")
        servicio_clean = cabecera.servicio.descripcion.replace(" ", "_")
        periodo_clean = cabecera.periodo.replace(" ", "_")
        filename = f"{empresa_clean}_{servicio_clean}_{periodo_clean}.xlsx"
        
        return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return f"Error generando Excel: {str(e)}", 500

@cuadrante_bp.route('/exportar/pdf/<int:cabecera_id>')
@login_required
def exportar_pdf(cabecera_id):
    """Generar PDF Vertical Premium con Resúmenes Mensuales."""
    cabecera = db.session.get(CuadranteCabecera, cabecera_id)
    if not cabecera:
        return "Cuadrante no encontrado", 404

    # 1. Preparar Estilos y Clases
    styles = getSampleStyleSheet()
    
    class VerticalText(Flowable):
        def __init__(self, text, font_size=7.5):
            Flowable.__init__(self)
            self.text = text
            self.font_size = font_size
            self.width = 16 # Ajustado para el ancho de 15.4cm
            self.height = 60

        def draw(self):
            canvas = self.canv
            canvas.saveState()
            canvas.setFont("Helvetica-Bold", self.font_size)
            canvas.setFillColor(colors.whitesmoke)
            canvas.rotate(90)
            canvas.drawString(8, -self.width + 5, self.text)
            canvas.restoreState()

        def wrap(self, availWidth, availHeight):
            return self.width, self.height

    # 2. Datos y Mapeos
    asignaciones = CuadranteAsignacion.query.filter_by(cabecera_id=cabecera_id).all()
    trabajadores_db = Trabajador.query.filter_by(empresa_id=cabecera.empresa_id, servicio_id=cabecera.servicio_id, activo=True).all()
    turnos_db = Turno.query.filter_by(empresa_id=cabecera.empresa_id, activo=True).all()
    
    t_nombres = {t.id: f"{t.nombre[0]}. {t.apellido1}" for t in trabajadores_db}
    turnos_map = {t.id: t.abreviacion for t in turnos_db}
    colores_map = {t.abreviacion: t.color for t in turnos_db if t.color}
    duraciones_map = {t.id: t.duracion_hrs for t in turnos_db}
    
    from app.models.core import Feriado
    import calendar
    _, last_day = calendar.monthrange(cabecera.anio, cabecera.mes)
    feriados = {f.fecha.day for f in Feriado.query.filter(db.extract('month', Feriado.fecha) == cabecera.mes, db.extract('year', Feriado.fecha) == cabecera.anio, Feriado.activo == True).all()}

    # 3. Preparar Matriz para ReportLab
    dias_nombres = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
    
    # Cabecera con Nombres Verticales para Trabajadores y Turnos
    header = ['Día'] + \
             [VerticalText(t_nombres[t.id], font_size=7) for t in trabajadores_db] + \
             [VerticalText(f"Total {t.nombre}", font_size=7) for t in turnos_db]
    table_data = [header]
    row_styles = []
    
    for d in range(1, last_day + 1):
        fecha_dt = datetime(cabecera.anio, cabecera.mes, d)
        nombre_dia = dias_nombres[fecha_dt.weekday()]
        is_holiday = d in feriados or fecha_dt.weekday() == 6
        
        row = [f"{nombre_dia} {d:02d}"]
        for t_obj in trabajadores_db:
            asig = next((a for a in asignaciones if a.fecha.day == d and a.trabajador_id == t_obj.id), None)
            abr = turnos_map.get(asig.turno_id, "L") if (asig and not asig.es_libre) else "L"
            row.append(abr)
            if abr != 'L' and abr in colores_map:
                row_styles.append(('BACKGROUND', (len(row)-1, d), (len(row)-1, d), colors.HexColor(colores_map[abr])))
        
        for t_tipo in turnos_db:
            count = sum(1 for a in asignaciones if a.fecha.day == d and not a.es_libre and turnos_map.get(a.turno_id) == t_tipo.abreviacion)
            row.append(str(count))

        table_data.append(row)
        if is_holiday:
            row_styles.append(('BACKGROUND', (0, d), (0, d), colors.red))
            row_styles.append(('TEXTCOLOR', (0, d), (0, d), colors.white))

    # Totales al final
    f_hrs, f_tns = ["HORAS"], ["TURNOS"]
    for t_obj in trabajadores_db:
        t_asigs = [a for a in asignaciones if a.trabajador_id == t_obj.id]
        hrs = sum(duraciones_map.get(a.turno_id, 0) for a in t_asigs if not a.es_libre)
        tns = sum(1 for a in t_asigs if not a.es_libre)
        f_hrs.append(str(int(hrs)))
        f_tns.append(str(tns))
    for _ in turnos_db:
        f_hrs.append(""); f_tns.append("")
    table_data.append(f_hrs); table_data.append(f_tns)

    # 4. Generar PDF con Medidas Exactas (15.4cm = 436.5pts)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, 
                            topMargin=15, bottomMargin=15, leftMargin=15, rightMargin=15,
                            title=f"Cuadrante_{cabecera.periodo}",
                            author=cabecera.empresa.razon_social)
    elements = []
    
    title_text = f"<b>Cuadrante Turnos {cabecera.empresa.razon_social} - {cabecera.servicio.descripcion} ({cabecera.periodo})</b>"
    t_title = styles['Normal']
    t_title.alignment = 1; t_title.fontSize = 12
    elements.append(Paragraph(title_text, t_title))
    elements.append(Spacer(1, 20))
    
    # Cálculos para 15.4 cm exactos
    total_w_pts = 436.5 
    col_day_w = 42
    remaining_w = total_w_pts - col_day_w
    col_other_w = remaining_w / (len(header) - 1)
    col_widths = [col_day_w] + [col_other_w] * (len(header) - 1)
    
    # Altura para que 31 días entren en una hoja (15.5pt es seguro)
    row_heights = [None] + [15.5] * (len(table_data) - 1)
    
    t = Table(table_data, repeatRows=1, colWidths=col_widths, rowHeights=row_heights, hAlign='CENTER')
    
    # Divisiones de sección (bordes más gruesos)
    idx_totals = len(trabajadores_db) + 1
    
    base_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1a2a3a")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), # Solo cabecera en blanco
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black), # El resto en negro
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, last_day+1), (0, -1), 'Helvetica-Bold'), # Negrita en totales
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.black),
        # Bordes de sección gruesos
        ('LINEBEFORE', (1, 0), (1, -1), 1.5, colors.black), 
        ('LINEBEFORE', (idx_totals, 0), (idx_totals, -1), 1.5, colors.black),
        ('LINEAFTER', (-1, 0), (-1, -1), 1.5, colors.black),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]
    t.setStyle(TableStyle(base_style + row_styles))
    elements.append(t)
    doc.build(elements)
    
    buffer.seek(0)
    # Nombre de archivo dinámico: Empresa_Servicio_Mes_Anio.pdf
    empresa_clean = cabecera.empresa.razon_social.replace(" ", "_")
    servicio_clean = cabecera.servicio.descripcion.replace(" ", "_")
    mes_str = dias_nombres[datetime(cabecera.anio, cabecera.mes, 1).weekday()] # Solo para el mes
    # Mejor usar el periodo formateado
    periodo_clean = cabecera.periodo.replace(" ", "_")
    filename = f"{empresa_clean}_{servicio_clean}_{periodo_clean}.pdf"
    
    return send_file(buffer, download_name=filename, as_attachment=True, mimetype='application/pdf')
